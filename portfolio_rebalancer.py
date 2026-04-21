import streamlit as st
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import re
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from datetime import datetime
import pdfplumber

# ---------------------------------------------------------------------------
# PAGE CONFIG & DARK THEME
# ---------------------------------------------------------------------------
st.set_page_config(page_title="Portfolio Rebalancer", layout="wide")

DARK_CSS = """
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
    .stApp { background-color: #000000; color: #ffffff; font-family: 'Inter', sans-serif; }
    header[data-testid="stHeader"] { background-color: #000000; }
    section[data-testid="stSidebar"] { background-color: #0a0a0a; border-right: 1px solid #1a1a1a; }
    section[data-testid="stSidebar"] * { color: #e0e0e0; }
    [data-testid="stMetric"] {
        background: #111111; border: 1px solid #1e1e1e;
        border-radius: 12px; padding: 16px 20px;
    }
    [data-testid="stMetricValue"] { color: #ffffff; font-weight: 600; }
    [data-testid="stMetricLabel"] { color: #8c8c8c; }
    .stDataFrame { border: 1px solid #1e1e1e; border-radius: 8px; overflow: hidden; }
    .stButton > button {
        background: #00c805; color: #000000; font-weight: 600;
        border: none; border-radius: 24px; padding: 12px 32px;
    }
    .stButton > button:hover { background: #00a804; color: #000000; }
    .stDownloadButton > button {
        background: #00c805; color: #000000; font-weight: 600;
        border: none; border-radius: 24px;
    }
    .buy-card {
        background: #071a07; border-left: 3px solid #00c805;
        padding: 16px 20px; border-radius: 8px; margin: 8px 0;
    }
    .sell-card {
        background: #1a0707; border-left: 3px solid #ff5000;
        padding: 16px 20px; border-radius: 8px; margin: 8px 0;
    }
    .hold-card {
        background: #111111; border-left: 3px solid #4a4a4a;
        padding: 16px 20px; border-radius: 8px; margin: 8px 0;
    }
    .info-card {
        background: #0a0a1a; border-left: 3px solid #3b82f6;
        padding: 16px 20px; border-radius: 8px; margin: 8px 0;
    }
    .warn-card {
        background: #1a1400; border-left: 3px solid #f59e0b;
        padding: 16px 20px; border-radius: 8px; margin: 8px 0;
    }
    .section-header {
        font-size: 1.3rem; font-weight: 600; margin: 40px 0 16px 0;
        padding-bottom: 10px; border-bottom: 1px solid #1e1e1e;
        color: #ffffff; letter-spacing: -0.02em;
    }
    .plan-step {
        background: #111111; border: 1px solid #1e1e1e;
        border-radius: 12px; padding: 20px; margin: 10px 0;
    }
    .green { color: #00c805; }
    .red { color: #ff5000; }
    .blue { color: #3b82f6; }
    .yellow { color: #f59e0b; }
    .muted { color: #6e6e6e; font-size: 0.92rem; }
    ::-webkit-scrollbar { width: 6px; }
    ::-webkit-scrollbar-track { background: #000000; }
    ::-webkit-scrollbar-thumb { background: #2a2a2a; border-radius: 3px; }
</style>
"""
st.markdown(DARK_CSS, unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# ASSET CLASS DEFINITIONS  (8 granular buckets)
# ---------------------------------------------------------------------------
ASSET_CLASSES = {
    "us_large_growth":   {"label": "US Large Cap Growth",       "color": "#00c805"},
    "us_broad":          {"label": "US Broad Market",           "color": "#3b82f6"},
    "us_dividend_value": {"label": "US Dividend / Value",       "color": "#22d3ee"},
    "intl_developed":    {"label": "International Developed",   "color": "#8b5cf6"},
    "intl_emerging":     {"label": "International Emerging",    "color": "#ec4899"},
    "us_core_bonds":     {"label": "US Core Bonds",             "color": "#f59e0b"},
    "treasury_tips":     {"label": "Treasury / TIPS",           "color": "#64748b"},
    "corporate_bonds":   {"label": "Corporate / HY Bonds",     "color": "#f97316"},
}

# Known ETF classifications
ETF_CLASS_MAP = {
    # US Large Cap Growth
    "QQQ": "us_large_growth", "SCHG": "us_large_growth", "VUG": "us_large_growth",
    "IWF": "us_large_growth", "MGK": "us_large_growth", "VONG": "us_large_growth",
    # US Broad Market
    "VOO": "us_broad", "VTI": "us_broad", "SPY": "us_broad", "IVV": "us_broad",
    "IWM": "us_broad", "DIA": "us_broad", "RSP": "us_broad", "ITOT": "us_broad",
    "SPTM": "us_broad", "SCHX": "us_broad", "SCHB": "us_broad",
    # US Dividend / Value
    "SCHD": "us_dividend_value", "VYM": "us_dividend_value", "VIG": "us_dividend_value",
    "DGRO": "us_dividend_value", "VTV": "us_dividend_value", "SCHV": "us_dividend_value",
    "HDV": "us_dividend_value", "USMV": "us_dividend_value", "QUAL": "us_dividend_value",
    "DVY": "us_dividend_value", "SDY": "us_dividend_value",
    # Sector ETFs -> us_broad (they're US stocks)
    "XLK": "us_large_growth", "XLF": "us_broad", "XLE": "us_broad",
    "XLV": "us_broad", "XLY": "us_broad", "XLP": "us_dividend_value",
    "XLI": "us_broad", "XLU": "us_dividend_value", "XLRE": "us_dividend_value",
    # International Developed
    "VEA": "intl_developed", "VXUS": "intl_developed", "EFA": "intl_developed",
    "IXUS": "intl_developed", "SCHF": "intl_developed", "SPDW": "intl_developed",
    # International Emerging
    "VWO": "intl_emerging", "EEM": "intl_emerging", "IEMG": "intl_emerging",
    # US Core Bonds
    "BND": "us_core_bonds", "AGG": "us_core_bonds", "SCHZ": "us_core_bonds",
    "BIV": "us_core_bonds", "BSV": "us_core_bonds", "MUB": "us_core_bonds",
    "BNDX": "us_core_bonds",
    # Treasury / TIPS
    "TLT": "treasury_tips", "IEF": "treasury_tips", "SHV": "treasury_tips",
    "SHY": "treasury_tips", "GOVT": "treasury_tips", "TIP": "treasury_tips",
    "SCHO": "treasury_tips", "SCHR": "treasury_tips", "VTIP": "treasury_tips",
    # Corporate / High-Yield
    "LQD": "corporate_bonds", "HYG": "corporate_bonds", "JNK": "corporate_bonds",
    "VCSH": "corporate_bonds", "VCIT": "corporate_bonds", "VCLT": "corporate_bonds",
    "EMB": "corporate_bonds", "BLV": "corporate_bonds",
}

# Sector mapping for yfinance sectors
SECTOR_MAP = {
    "Technology": "tech", "Communication Services": "tech",
    "Consumer Cyclical": "consumer", "Consumer Defensive": "consumer",
    "Healthcare": "healthcare", "Financial Services": "financials",
    "Industrials": "industrials", "Energy": "energy",
    "Utilities": "utilities", "Real Estate": "real_estate",
    "Basic Materials": "materials",
}

SECTOR_LABELS = {
    "tech": "Technology", "consumer": "Consumer", "healthcare": "Healthcare",
    "financials": "Financials", "industrials": "Industrials", "energy": "Energy",
    "utilities": "Utilities", "real_estate": "Real Estate", "materials": "Materials",
    "etf": "ETF / Fund", "unknown": "Other",
}

# ---------------------------------------------------------------------------
# ADVISOR AGENTS  — 3 distinct strategy profiles
# ---------------------------------------------------------------------------
# Each agent has allocation targets across all 8 asset classes,
# plus modifiers based on time horizon, age, and goal.

ADVISOR_AGENTS = {
    "Conservative": {
        "name": "Capital Shield Advisor",
        "philosophy": "Protect what you have. This strategy prioritizes capital preservation "
                      "and steady income over growth. It favors high-quality bonds, dividend-paying "
                      "stocks, and minimal exposure to volatile growth sectors.",
        "base_allocation": {
            "us_large_growth":   0.00,
            "us_broad":          0.10,
            "us_dividend_value": 0.15,
            "intl_developed":    0.05,
            "intl_emerging":     0.00,
            "us_core_bonds":     0.35,
            "treasury_tips":     0.25,
            "corporate_bonds":   0.10,
        },
        "etf_picks": {
            "us_broad":          {"ticker": "VOO",  "name": "Vanguard S&P 500 ETF",               "why": "Core blue-chip exposure with low volatility relative to growth funds"},
            "us_dividend_value": {"ticker": "SCHD", "name": "Schwab US Dividend Equity ETF",       "why": "High-quality dividend payers for reliable income and downside protection"},
            "intl_developed":    {"ticker": "VEA",  "name": "Vanguard FTSE Developed Markets ETF", "why": "Modest international diversification across stable economies"},
            "us_core_bonds":     {"ticker": "BND",  "name": "Vanguard Total Bond Market ETF",      "why": "Broad investment-grade bond exposure for core stability"},
            "treasury_tips":     {"ticker": "TIP",  "name": "iShares TIPS Bond ETF",               "why": "Inflation-protected Treasuries to preserve purchasing power"},
            "corporate_bonds":   {"ticker": "VCSH", "name": "Vanguard Short-Term Corp Bond ETF",   "why": "Short duration minimizes interest-rate risk while earning a yield premium"},
        },
        "sector_guidance": {
            "tech": "Limit technology exposure. Individual tech stocks carry concentration risk that conflicts with a preservation strategy.",
            "energy": "Energy can be volatile. If holding energy, prefer pipeline companies or utilities with steady dividends.",
            "healthcare": "Defensive healthcare holdings (like JNJ, PFE) can fit a conservative portfolio for their dividend stability.",
        },
    },
    "Balanced": {
        "name": "Growth & Income Advisor",
        "philosophy": "Grow steadily while managing risk. This strategy blends broad market growth "
                      "with dividend income and a meaningful bond allocation. It aims for consistent "
                      "returns across market cycles without extreme concentration.",
        "base_allocation": {
            "us_large_growth":   0.10,
            "us_broad":          0.25,
            "us_dividend_value": 0.10,
            "intl_developed":    0.08,
            "intl_emerging":     0.02,
            "us_core_bonds":     0.25,
            "treasury_tips":     0.10,
            "corporate_bonds":   0.10,
        },
        "etf_picks": {
            "us_large_growth":   {"ticker": "QQQ",  "name": "Invesco QQQ Trust (Nasdaq-100)",      "why": "Measured growth allocation for long-term capital appreciation"},
            "us_broad":          {"ticker": "VOO",  "name": "Vanguard S&P 500 ETF",               "why": "Core US equity exposure tracking the benchmark index"},
            "us_dividend_value": {"ticker": "SCHD", "name": "Schwab US Dividend Equity ETF",       "why": "Dividend income adds a return floor during market downturns"},
            "intl_developed":    {"ticker": "VEA",  "name": "Vanguard FTSE Developed Markets ETF", "why": "Geographic diversification reduces US-only concentration risk"},
            "intl_emerging":     {"ticker": "VWO",  "name": "Vanguard FTSE Emerging Markets ETF",  "why": "Small emerging markets allocation for long-term growth potential"},
            "us_core_bonds":     {"ticker": "BND",  "name": "Vanguard Total Bond Market ETF",      "why": "Investment-grade bonds anchor portfolio stability"},
            "treasury_tips":     {"ticker": "TIP",  "name": "iShares TIPS Bond ETF",               "why": "Inflation protection for the fixed-income sleeve"},
            "corporate_bonds":   {"ticker": "VCIT", "name": "Vanguard Interm-Term Corp Bond ETF",  "why": "Corporate bonds offer a yield premium over Treasuries"},
        },
        "sector_guidance": {
            "tech": "Technology is fine at up to 25% of your portfolio. Beyond that, consider trimming and diversifying.",
            "energy": "A small energy allocation (5-10%) provides inflation hedging and diversification.",
            "healthcare": "Healthcare is a solid balanced-portfolio sector. Stable demand regardless of economic conditions.",
        },
    },
    "Aggressive": {
        "name": "Maximum Growth Advisor",
        "philosophy": "Maximize long-term wealth creation. This strategy accepts higher short-term "
                      "volatility in exchange for superior long-term returns. It favors growth equities, "
                      "broad market exposure, and minimal bonds — just enough to rebalance during downturns.",
        "base_allocation": {
            "us_large_growth":   0.25,
            "us_broad":          0.30,
            "us_dividend_value": 0.05,
            "intl_developed":    0.10,
            "intl_emerging":     0.05,
            "us_core_bonds":     0.10,
            "treasury_tips":     0.05,
            "corporate_bonds":   0.10,
        },
        "etf_picks": {
            "us_large_growth":   {"ticker": "QQQ",  "name": "Invesco QQQ Trust (Nasdaq-100)",      "why": "Heavy growth tilt for maximum capital appreciation over time"},
            "us_broad":          {"ticker": "VTI",  "name": "Vanguard Total Stock Market ETF",     "why": "Full US market breadth including mid and small caps for growth"},
            "us_dividend_value": {"ticker": "SCHD", "name": "Schwab US Dividend Equity ETF",       "why": "Small dividend anchor for downside protection during corrections"},
            "intl_developed":    {"ticker": "VEA",  "name": "Vanguard FTSE Developed Markets ETF", "why": "International diversification captures non-US growth cycles"},
            "intl_emerging":     {"ticker": "VWO",  "name": "Vanguard FTSE Emerging Markets ETF",  "why": "Emerging market exposure for higher long-term growth trajectory"},
            "us_core_bonds":     {"ticker": "BND",  "name": "Vanguard Total Bond Market ETF",      "why": "Minimal bond allocation provides dry powder to buy dips"},
            "treasury_tips":     {"ticker": "SHV",  "name": "iShares Short Treasury Bond ETF",     "why": "Near-cash safety for tactical rebalancing opportunities"},
            "corporate_bonds":   {"ticker": "VCIT", "name": "Vanguard Interm-Term Corp Bond ETF",  "why": "Higher-yielding corporates maximize bond-sleeve returns"},
        },
        "sector_guidance": {
            "tech": "Growth strategies can tolerate higher tech exposure (up to 35%), but watch single-stock concentration.",
            "energy": "Energy is cyclical — fine as a small allocation but don't overweight in a growth portfolio.",
            "healthcare": "Biotech and healthcare innovation fits well in an aggressive strategy.",
        },
    },
}


def adjust_allocation_for_profile(base_alloc, time_horizon, age_range, goal, intl_pref):
    """Modify the base allocation based on the investor's full profile.
    Returns a new allocation dict and a list of adjustment reasons."""
    alloc = dict(base_alloc)
    reasons = []

    # --- Time horizon adjustments ---
    if time_horizon == "Under 3 years":
        # Shift heavily toward bonds regardless of risk tolerance
        shift = 0.15
        alloc["us_core_bonds"] = min(alloc["us_core_bonds"] + shift * 0.5, 0.45)
        alloc["treasury_tips"] = min(alloc["treasury_tips"] + shift * 0.3, 0.30)
        alloc["corporate_bonds"] = min(alloc["corporate_bonds"] + shift * 0.2, 0.20)
        alloc["us_large_growth"] = max(alloc["us_large_growth"] - shift * 0.5, 0.0)
        alloc["us_broad"] = max(alloc["us_broad"] - shift * 0.3, 0.05)
        alloc["intl_emerging"] = max(alloc["intl_emerging"] - shift * 0.2, 0.0)
        reasons.append("Short time horizon (under 3 years): increased bond allocation and reduced volatile equity exposure to protect against short-term drawdowns.")

    elif time_horizon == "10+ years":
        shift = 0.05
        alloc["us_broad"] = min(alloc["us_broad"] + shift * 0.5, 0.40)
        alloc["us_large_growth"] = min(alloc["us_large_growth"] + shift * 0.5, 0.35)
        alloc["us_core_bonds"] = max(alloc["us_core_bonds"] - shift * 0.5, 0.05)
        alloc["treasury_tips"] = max(alloc["treasury_tips"] - shift * 0.5, 0.02)
        reasons.append("Long time horizon (10+ years): slightly increased equity allocation to capture long-term compounding.")

    # --- Age adjustments ---
    if age_range == "60+":
        shift = 0.10
        alloc["us_dividend_value"] = min(alloc["us_dividend_value"] + shift * 0.4, 0.25)
        alloc["us_core_bonds"] = min(alloc["us_core_bonds"] + shift * 0.3, 0.45)
        alloc["treasury_tips"] = min(alloc["treasury_tips"] + shift * 0.3, 0.30)
        alloc["us_large_growth"] = max(alloc["us_large_growth"] - shift * 0.5, 0.0)
        alloc["intl_emerging"] = max(alloc["intl_emerging"] - shift * 0.3, 0.0)
        alloc["us_broad"] = max(alloc["us_broad"] - shift * 0.2, 0.05)
        reasons.append("Age 60+: shifted toward income-producing assets and reduced volatile growth exposure to protect against sequence-of-returns risk.")
    elif age_range == "50 - 59":
        shift = 0.05
        alloc["us_dividend_value"] = min(alloc["us_dividend_value"] + shift * 0.5, 0.20)
        alloc["us_core_bonds"] = min(alloc["us_core_bonds"] + shift * 0.5, 0.40)
        alloc["us_large_growth"] = max(alloc["us_large_growth"] - shift * 0.5, 0.0)
        alloc["intl_emerging"] = max(alloc["intl_emerging"] - shift * 0.5, 0.0)
        reasons.append("Age 50-59: moderately increased income and bond allocation as retirement approaches.")
    elif age_range == "18 - 29":
        shift = 0.03
        alloc["us_large_growth"] = min(alloc["us_large_growth"] + shift, 0.35)
        alloc["us_core_bonds"] = max(alloc["us_core_bonds"] - shift, 0.05)
        reasons.append("Age 18-29: long runway to retirement allows slightly more growth-oriented allocation.")

    # --- Goal adjustments ---
    if goal == "Generate income":
        shift = 0.08
        alloc["us_dividend_value"] = min(alloc["us_dividend_value"] + shift * 0.5, 0.25)
        alloc["corporate_bonds"] = min(alloc["corporate_bonds"] + shift * 0.3, 0.20)
        alloc["us_core_bonds"] = min(alloc["us_core_bonds"] + shift * 0.2, 0.40)
        alloc["us_large_growth"] = max(alloc["us_large_growth"] - shift * 0.5, 0.0)
        alloc["intl_emerging"] = max(alloc["intl_emerging"] - shift * 0.3, 0.0)
        alloc["us_broad"] = max(alloc["us_broad"] - shift * 0.2, 0.05)
        reasons.append("Income goal: overweighted dividend stocks and higher-yielding bonds to maximize portfolio income.")
    elif goal == "Preserve capital":
        shift = 0.08
        alloc["treasury_tips"] = min(alloc["treasury_tips"] + shift * 0.5, 0.35)
        alloc["us_core_bonds"] = min(alloc["us_core_bonds"] + shift * 0.3, 0.45)
        alloc["us_dividend_value"] = min(alloc["us_dividend_value"] + shift * 0.2, 0.20)
        alloc["us_large_growth"] = max(alloc["us_large_growth"] - shift * 0.5, 0.0)
        alloc["intl_emerging"] = max(alloc["intl_emerging"] - shift * 0.3, 0.0)
        alloc["us_broad"] = max(alloc["us_broad"] - shift * 0.2, 0.05)
        reasons.append("Preservation goal: maximized treasury and bond exposure to protect capital.")

    # --- International preference ---
    if intl_pref == "US only":
        intl_total = alloc["intl_developed"] + alloc["intl_emerging"]
        alloc["us_broad"] += intl_total * 0.6
        alloc["us_dividend_value"] += intl_total * 0.4
        alloc["intl_developed"] = 0.0
        alloc["intl_emerging"] = 0.0
        reasons.append("US only preference: reallocated international exposure to domestic equity.")

    # Normalize to sum to 1.0
    total = sum(alloc.values())
    if total > 0:
        alloc = {k: v / total for k, v in alloc.items()}

    return alloc, reasons


# ---------------------------------------------------------------------------
# HOLDING ANALYSIS — uses yfinance for real sector/type data
# ---------------------------------------------------------------------------
@st.cache_data(ttl=300, show_spinner=False)
def get_holding_info(ticker):
    """Fetch sector, name, quote type, and price for a holding via yfinance."""
    try:
        t = yf.Ticker(ticker)
        info = t.info
        fast = t.fast_info
        return {
            "name": info.get("shortName", info.get("longName", ticker)),
            "sector": info.get("sector", ""),
            "industry": info.get("industry", ""),
            "quote_type": info.get("quoteType", ""),
            "price": getattr(fast, "last_price", 0) or 0,
            "market_cap": info.get("marketCap", 0),
            "dividend_yield": info.get("dividendYield", 0) or 0,
            "pe_ratio": info.get("forwardPE", info.get("trailingPE", 0)) or 0,
        }
    except Exception:
        return {
            "name": ticker, "sector": "", "industry": "", "quote_type": "",
            "price": 0, "market_cap": 0, "dividend_yield": 0, "pe_ratio": 0,
        }


def classify_holding_detailed(ticker, info):
    """Classify a holding into one of the 8 asset classes using ETF map + yfinance data."""
    t = ticker.upper()
    if t in ETF_CLASS_MAP:
        return ETF_CLASS_MAP[t]

    qtype = info.get("quote_type", "").upper()
    name = info.get("name", "").lower()

    # Bond funds / fixed income
    if any(w in name for w in ["bond", "treasury", "fixed income", "income fund", "debt"]):
        if any(w in name for w in ["corporate", "high yield", "high-yield"]):
            return "corporate_bonds"
        if any(w in name for w in ["treasury", "tips", "inflation"]):
            return "treasury_tips"
        return "us_core_bonds"

    # International
    if any(w in name for w in ["international", "emerging", "global ex-us", "foreign"]):
        if any(w in name for w in ["emerging", "em "]):
            return "intl_emerging"
        return "intl_developed"

    # Dividend / Value
    if any(w in name for w in ["dividend", "value", "income equity"]):
        return "us_dividend_value"

    # Growth
    if any(w in name for w in ["growth", "nasdaq", "innovation", "technology"]):
        return "us_large_growth"

    # Individual stocks — classify by sector
    sector = info.get("sector", "")
    if sector in ("Technology", "Communication Services"):
        return "us_large_growth"
    else:
        return "us_broad"  # default for US individual stocks


def get_sector_for_holding(info):
    """Get the simplified sector label for a holding."""
    sector = info.get("sector", "")
    qtype = info.get("quote_type", "").upper()
    if qtype in ("ETF", "MUTUALFUND"):
        return "etf"
    return SECTOR_MAP.get(sector, "unknown")


def get_live_price(ticker):
    """Fetch live price for a ticker."""
    try:
        return round(yf.Ticker(ticker).fast_info.last_price, 2)
    except Exception:
        return None


# ---------------------------------------------------------------------------
# BROKERAGE PARSERS
# ---------------------------------------------------------------------------
BROKERAGE_FORMATS = {
    "Generic (Ticker, Shares, Price)": {
        "ticker_cols": ["TICKER", "SYMBOL", "STOCK"],
        "shares_cols": ["SHARES", "QUANTITY", "QTY"],
        "price_cols": ["PRICE", "CURRENT PRICE", "COST", "LAST PRICE"],
    },
    "Robinhood": {
        "ticker_cols": ["SYMBOL", "INSTRUMENT"],
        "shares_cols": ["QUANTITY", "SHARES"],
        "price_cols": ["AVERAGE COST", "CURRENT PRICE", "LAST PRICE"],
    },
    "Fidelity": {
        "ticker_cols": ["SYMBOL"],
        "shares_cols": ["QUANTITY"],
        "price_cols": ["LAST PRICE", "CURRENT VALUE"],
    },
    "Charles Schwab": {
        "ticker_cols": ["SYMBOL"],
        "shares_cols": ["QUANTITY"],
        "price_cols": ["PRICE", "MARKET VALUE"],
    },
    "E*Trade": {
        "ticker_cols": ["SYMBOL"],
        "shares_cols": ["QUANTITY"],
        "price_cols": ["PRICE PAID", "LAST PRICE"],
    },
    "TD Ameritrade / Schwab": {
        "ticker_cols": ["SYMBOL"],
        "shares_cols": ["QUANTITY", "QTY"],
        "price_cols": ["LAST", "MARK", "PRICE"],
    },
}


def find_column(df, candidates):
    for col in df.columns:
        if col.strip().upper() in [c.upper() for c in candidates]:
            return col
    return None


def parse_brokerage_csv(df, fmt_key):
    fmt = BROKERAGE_FORMATS[fmt_key]
    ticker_col = find_column(df, fmt["ticker_cols"])
    shares_col = find_column(df, fmt["shares_cols"])
    price_col = find_column(df, fmt["price_cols"])
    data = {}
    if not ticker_col or not shares_col:
        return data, "Could not find required columns (Ticker and Shares) in your CSV."
    for _, row in df.iterrows():
        try:
            ticker = str(row[ticker_col]).strip().upper()
            if not ticker or ticker in ("", "NAN", "CASH", "PENDING", "--"):
                continue
            ticker = ticker.replace("$", "").strip()
            shares = float(str(row[shares_col]).replace(",", "").replace("$", ""))
            if shares <= 0:
                continue
            if price_col and pd.notna(row[price_col]):
                price_str = str(row[price_col]).replace(",", "").replace("$", "")
                price = float(price_str) if price_str else 0
            else:
                price = 0
            if price <= 0 or price > 50000:
                live = get_live_price(ticker)
                if live:
                    price = live
            data[ticker] = {"shares": shares, "price": price}
        except (ValueError, TypeError):
            continue
    return data, None


def parse_robinhood_pdf(pdf_file):
    """Parse Robinhood monthly statement PDF."""
    data = {}
    ticker_pattern = re.compile(r'^([A-Z]{1,5})$')
    skip_words = {"CASH", "TOTAL", "PAGE", "NAN", "MARGIN", "USD", "SWEEP", "FDIC"}
    try:
        with pdfplumber.open(pdf_file) as pdf:
            all_text = ""
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    all_text += text + "\n"
            lines = all_text.split("\n")
            in_portfolio = False
            for line in lines:
                stripped = line.strip()
                if "Portfolio Summary" in stripped or "Securities Held in Account" in stripped:
                    in_portfolio = True
                    continue
                if in_portfolio and any(kw in stripped for kw in [
                    "Account Activity", "Executed Trades", "Deposit Sweep",
                    "Brokerage-held Cash Activity", "Options Cash Collateral",
                    "Important Information"]):
                    in_portfolio = False
                    continue
                if not in_portfolio or not stripped:
                    continue
                if "Estimated Yield" in stripped or "Sym/Cusip" in stripped or "Securities Held" in stripped:
                    continue
                if any(kw in stripped for kw in ["Total Securities", "Total Priced", "Brokerage Cash Balance", "Deposit Sweep Balance"]):
                    continue
                parts = stripped.split()
                if len(parts) < 4:
                    continue
                candidate = parts[0].strip()
                if not ticker_pattern.match(candidate) or candidate in skip_words:
                    continue
                if "Margin" not in parts:
                    continue
                mi = parts.index("Margin")
                try:
                    qty = float(parts[mi + 1].replace(",", ""))
                    price = float(parts[mi + 2].replace("$", "").replace(",", ""))
                    if qty > 0 and price > 0:
                        data[candidate] = {"shares": qty, "price": price}
                except (IndexError, ValueError):
                    continue
    except Exception as e:
        return data, f"Error reading Robinhood PDF: {e}"
    return data, None


def parse_pdf_statement(pdf_file):
    """Parse brokerage PDF — tries Robinhood first, then generic."""
    try:
        pdf_file.seek(0)
        with pdfplumber.open(pdf_file) as pdf:
            first_text = pdf.pages[0].extract_text() or "" if pdf.pages else ""
        if "robinhood" in first_text.lower():
            pdf_file.seek(0)
            data, err = parse_robinhood_pdf(pdf_file)
            if data:
                return data, err
        # Generic fallback (table-based)
        pdf_file.seek(0)
        with pdfplumber.open(pdf_file) as pdf:
            ticker_pat = re.compile(r'\b([A-Z]{1,5})\b')
            num_pat = re.compile(r'[\d,]+\.?\d*')
            data = {}
            for page in pdf.pages:
                for table in (page.extract_tables() or []):
                    if not table or len(table) < 2:
                        continue
                    header = [str(c).strip().upper() if c else "" for c in table[0]]
                    ti = si = pi = vi = None
                    for i, h in enumerate(header):
                        if any(k in h for k in ["SYMBOL", "TICKER", "SYM"]):
                            ti = i
                        if any(k in h for k in ["SHARES", "QUANTITY", "QTY"]):
                            si = i
                        if any(k in h for k in ["PRICE", "LAST", "COST"]):
                            pi = i
                        if any(k in h for k in ["VALUE", "MKT VALUE"]):
                            vi = i
                    if ti is not None and (si is not None or vi is not None):
                        for row in table[1:]:
                            try:
                                cell = str(row[ti]).strip().upper()
                                tks = ticker_pat.findall(cell)
                                if not tks:
                                    continue
                                tk = tks[0]
                                if tk in ("TOTAL", "CASH", "NAN"):
                                    continue
                                shares = float(num_pat.findall(str(row[si]).replace(",", ""))[0]) if si and row[si] else 0
                                price = float(num_pat.findall(str(row[pi]).replace(",", "").replace("$", ""))[0]) if pi and row[pi] else 0
                                if price == 0 and vi and row[vi] and shares > 0:
                                    price = float(num_pat.findall(str(row[vi]).replace(",", "").replace("$", ""))[0]) / shares
                                if shares > 0 and price <= 0:
                                    live = get_live_price(tk)
                                    if live:
                                        price = live
                                if tk and shares > 0 and price > 0:
                                    data[tk] = {"shares": shares, "price": price}
                            except Exception:
                                continue
        if not data:
            return data, "Could not find holdings in this PDF. Try CSV or manual entry."
        return data, None
    except Exception as e:
        return {}, f"Error reading PDF: {e}"


# ---------------------------------------------------------------------------
# CHART HELPERS
# ---------------------------------------------------------------------------
def dark_pie_chart(values, labels, title, colors=None):
    fig, ax = plt.subplots(figsize=(6, 5))
    fig.patch.set_facecolor('#000000')
    ax.set_facecolor('#000000')
    if colors is None:
        colors = ['#00c805', '#3b82f6', '#f59e0b', '#ff5000', '#8b5cf6',
                  '#ec4899', '#14b8a6', '#f97316', '#6366f1', '#84cc16']
    wedges, texts, autotexts = ax.pie(
        values, labels=None, autopct='%1.1f%%', startangle=90,
        colors=colors[:len(values)],
        textprops={'color': '#ffffff', 'fontsize': 10},
        wedgeprops={'linewidth': 1, 'edgecolor': '#1e1e1e'},
    )
    ax.legend(labels, loc='lower center', bbox_to_anchor=(0.5, -0.18),
              ncol=min(3, len(labels)), fontsize=8,
              facecolor='#111111', edgecolor='#1e1e1e', labelcolor='#e0e0e0')
    ax.set_title(title, color='#ffffff', fontsize=13, fontweight='bold', pad=12)
    plt.tight_layout()
    return fig


def dark_bar_chart(categories, current_pcts, target_pcts, title):
    fig, ax = plt.subplots(figsize=(8, 5))
    fig.patch.set_facecolor('#000000')
    ax.set_facecolor('#000000')
    x = range(len(categories))
    w = 0.35
    bars1 = ax.bar([i - w/2 for i in x], current_pcts, w, label='Current', color='#ff5000', edgecolor='#1e1e1e')
    bars2 = ax.bar([i + w/2 for i in x], target_pcts, w, label='Target', color='#00c805', edgecolor='#1e1e1e')
    ax.set_xticks(list(x))
    ax.set_xticklabels(categories, rotation=30, ha='right', fontsize=8, color='#e0e0e0')
    ax.set_ylabel('%', color='#e0e0e0')
    ax.tick_params(colors='#6e6e6e')
    ax.spines['bottom'].set_color('#1e1e1e')
    ax.spines['left'].set_color('#1e1e1e')
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.legend(facecolor='#111111', edgecolor='#1e1e1e', labelcolor='#e0e0e0')
    ax.set_title(title, color='#ffffff', fontsize=13, fontweight='bold', pad=12)
    plt.tight_layout()
    return fig


# =========================================================================
#  STREAMLIT APP
# =========================================================================
st.markdown("<h1 style='text-align:center; color:#ffffff; font-weight:700; letter-spacing:-0.03em;'>Portfolio Rebalancer</h1>", unsafe_allow_html=True)
st.markdown("<p style='text-align:center; color:#6e6e6e; font-size:1rem;'>Intelligent rebalancing powered by data-driven strategy</p>", unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# SIDEBAR — ONBOARDING
# ---------------------------------------------------------------------------
st.sidebar.markdown("### Investor Profile")

risk_level = st.sidebar.radio(
    "Risk tolerance",
    ["Conservative", "Balanced", "Aggressive"],
    index=1,
    help="How comfortable are you with short-term losses for long-term growth?",
)
time_horizon = st.sidebar.selectbox("Investment time horizon", ["Under 3 years", "3 - 10 years", "10+ years"], index=1)
goal = st.sidebar.selectbox("Primary investment goal", ["Grow wealth over time", "Generate income", "Preserve capital"])
age_range = st.sidebar.selectbox("Age range", ["18 - 29", "30 - 39", "40 - 49", "50 - 59", "60+"])
intl_pref = st.sidebar.radio("International exposure", ["US only", "Include international"])

# --- Compute advisor allocation ---
agent = ADVISOR_AGENTS[risk_level]
target_alloc, profile_reasons = adjust_allocation_for_profile(
    agent["base_allocation"], time_horizon, age_range, goal, intl_pref
)

if time_horizon == "Under 3 years" and risk_level == "Aggressive":
    st.sidebar.warning("With under 3 years, aggressive carries significant short-term risk. Your allocation has been adjusted.")
if age_range == "60+" and risk_level == "Aggressive":
    st.sidebar.warning("Near retirement, aggressive allocation may expose you to sequence-of-returns risk. Your allocation has been adjusted.")

st.sidebar.markdown("---")
st.sidebar.markdown("### Portfolio Input")
input_method = st.sidebar.radio("How would you like to enter your portfolio?", ["Manual Entry", "Upload CSV / PDF"])

portfolio_data = {}

if input_method == "Manual Entry":
    num_holdings = st.sidebar.number_input("Number of holdings", min_value=1, max_value=30, value=3)
    for i in range(int(num_holdings)):
        st.sidebar.markdown(f"**Holding {i+1}**")
        search_q = st.sidebar.text_input("Company or ticker", key=f"search_{i}", placeholder="e.g. Apple or AAPL")
        ticker = ""
        price = 0.0
        if search_q:
            try:
                results = yf.Search(search_q, max_results=5).quotes
                if results:
                    options = {f"{r.get('shortname', r.get('symbol', ''))} ({r['symbol']})": r['symbol'] for r in results if 'symbol' in r}
                    if options:
                        selected = st.sidebar.selectbox("Select", list(options.keys()), key=f"sel_{i}")
                        ticker = options[selected]
                        live = get_live_price(ticker)
                        if live:
                            price = live
                            st.sidebar.markdown(f"<span class='green'>Current price: ${price:,.2f}</span>", unsafe_allow_html=True)
                        else:
                            price = st.sidebar.number_input("Price", min_value=0.0, key=f"price_{i}")
                    else:
                        ticker = search_q.upper()
                        price = st.sidebar.number_input("Price", min_value=0.0, key=f"price_{i}")
                else:
                    ticker = search_q.upper()
                    price = st.sidebar.number_input("Price", min_value=0.0, key=f"price_{i}")
            except Exception:
                ticker = search_q.upper()
                price = st.sidebar.number_input("Price", min_value=0.0, key=f"price_{i}")
        shares = st.sidebar.number_input("Shares", min_value=0.0, key=f"shares_{i}")
        if ticker and shares > 0 and price > 0:
            portfolio_data[ticker] = {"shares": shares, "price": price}
else:
    brokerage = st.sidebar.selectbox("Brokerage format", list(BROKERAGE_FORMATS.keys()))
    uploaded = st.sidebar.file_uploader("Upload positions file", type=["csv", "pdf"])
    if uploaded:
        fname = uploaded.name.lower()
        try:
            if fname.endswith(".pdf"):
                parsed, err = parse_pdf_statement(uploaded)
            else:
                df_raw = pd.read_csv(uploaded)
                parsed, err = parse_brokerage_csv(df_raw, brokerage)
            if err:
                st.sidebar.error(err)
            elif parsed:
                portfolio_data = parsed
                st.sidebar.success(f"Loaded {len(parsed)} holdings")
            else:
                st.sidebar.warning("No holdings found. Try another format or manual entry.")
        except Exception as e:
            st.sidebar.error(f"Error: {e}")

# =========================================================================
#  MAIN DASHBOARD
# =========================================================================
if portfolio_data:
    total_value = sum(d["shares"] * d["price"] for d in portfolio_data.values())
    if total_value <= 0:
        st.error("Portfolio total value is $0. Check your inputs.")
        st.stop()

    # --- Fetch info for all holdings ---
    with st.spinner("Analyzing your holdings..."):
        holdings = {}
        for ticker, d in portfolio_data.items():
            info = get_holding_info(ticker)
            val = d["shares"] * d["price"]
            asset_class = classify_holding_detailed(ticker, info)
            sector = get_sector_for_holding(info)
            holdings[ticker] = {
                "shares": d["shares"],
                "price": d["price"],
                "value": val,
                "pct": (val / total_value) * 100,
                "asset_class": asset_class,
                "sector": sector,
                "info": info,
            }

    # Compute current allocation by asset class
    current_by_class = {}
    for cls in ASSET_CLASSES:
        current_by_class[cls] = sum(h["value"] for h in holdings.values() if h["asset_class"] == cls)
    current_pct_by_class = {cls: (v / total_value) * 100 for cls, v in current_by_class.items()}

    # Compute sector exposure
    sector_exposure = {}
    for h in holdings.values():
        s = h["sector"]
        sector_exposure[s] = sector_exposure.get(s, 0) + h["pct"]

    # Target values
    target_by_class = {cls: target_alloc[cls] * total_value for cls in ASSET_CLASSES}
    target_pct_by_class = {cls: target_alloc[cls] * 100 for cls in ASSET_CLASSES}

    # Diffs
    diff_by_class = {cls: target_by_class[cls] - current_by_class.get(cls, 0) for cls in ASSET_CLASSES}

    total_stocks_pct = sum(current_pct_by_class.get(c, 0) for c in ["us_large_growth", "us_broad", "us_dividend_value", "intl_developed", "intl_emerging"])
    total_bonds_pct = sum(current_pct_by_class.get(c, 0) for c in ["us_core_bonds", "treasury_tips", "corporate_bonds"])

    # ==================================================================
    # SECTION 1: YOUR CURRENT PORTFOLIO
    # ==================================================================
    st.markdown("<div class='section-header'>Your Current Portfolio</div>", unsafe_allow_html=True)

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total Value", f"${total_value:,.2f}")
    m2.metric("Holdings", len(holdings))
    m3.metric("Equities", f"{total_stocks_pct:.1f}%")
    m4.metric("Fixed Income", f"{total_bonds_pct:.1f}%")

    col_t, col_c = st.columns([3, 2])
    with col_t:
        rows = []
        for t, h in sorted(holdings.items(), key=lambda x: -x[1]["value"]):
            rows.append({
                "Ticker": t,
                "Name": h["info"]["name"][:30],
                "Sector": SECTOR_LABELS.get(h["sector"], h["sector"]),
                "Class": ASSET_CLASSES.get(h["asset_class"], {}).get("label", h["asset_class"]),
                "Shares": round(h["shares"], 4),
                "Price": f"${h['price']:,.2f}",
                "Value": f"${h['value']:,.2f}",
                "Weight": f"{h['pct']:.1f}%",
            })
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    with col_c:
        # Pie by asset class
        ac_vals = [current_pct_by_class.get(c, 0) for c in ASSET_CLASSES if current_pct_by_class.get(c, 0) > 0]
        ac_labels = [ASSET_CLASSES[c]["label"] for c in ASSET_CLASSES if current_pct_by_class.get(c, 0) > 0]
        ac_colors = [ASSET_CLASSES[c]["color"] for c in ASSET_CLASSES if current_pct_by_class.get(c, 0) > 0]
        if ac_vals:
            fig1 = dark_pie_chart(ac_vals, ac_labels, "Current Allocation by Asset Class", ac_colors)
            st.pyplot(fig1)
            plt.close(fig1)

    # ==================================================================
    # SECTION 2: PORTFOLIO ANALYSIS
    # ==================================================================
    st.markdown("<div class='section-header'>Portfolio Analysis</div>", unsafe_allow_html=True)

    warnings = []
    insights = []

    # Concentration
    for t, h in holdings.items():
        if h["pct"] > 20:
            warnings.append(f"**{t}** ({h['info']['name'][:25]}) is {h['pct']:.1f}% of your portfolio. "
                            f"Single-stock concentration above 20% significantly increases risk. Consider trimming.")

    # Sector concentration
    for s, pct in sorted(sector_exposure.items(), key=lambda x: -x[1]):
        if s not in ("etf", "unknown") and pct > 30:
            warnings.append(f"**{SECTOR_LABELS.get(s, s)}** sector is {pct:.1f}% of your portfolio. "
                            f"Heavy sector concentration means a downturn in that industry could significantly impact your returns.")

    # Diversification
    if len(holdings) < 5:
        warnings.append("Your portfolio has fewer than 5 holdings. Limited diversification increases volatility.")
    if total_bonds_pct == 0 and risk_level != "Aggressive":
        warnings.append("You have **zero fixed-income exposure**. Even growth-oriented investors typically hold some bonds as a shock absorber.")

    # Positive observations
    if len(holdings) >= 10:
        insights.append("Good diversification with " + str(len(holdings)) + " holdings across your portfolio.")
    if total_bonds_pct > 0 and total_stocks_pct > 0:
        insights.append(f"You have a {total_stocks_pct:.0f}/{total_bonds_pct:.0f} stock/bond split, providing some balance.")

    # Sector guidance from the advisor
    for s, pct in sector_exposure.items():
        if s in agent["sector_guidance"] and pct > 15:
            insights.append(f"**{SECTOR_LABELS.get(s, s)}** ({pct:.1f}%): {agent['sector_guidance'][s]}")

    if warnings:
        for w in warnings:
            st.markdown(f"<div class='warn-card'>{w}</div>", unsafe_allow_html=True)
    if insights:
        for ins in insights:
            st.markdown(f"<div class='info-card'>{ins}</div>", unsafe_allow_html=True)

    # Sector breakdown bar chart
    if sector_exposure:
        sec_col1, sec_col2 = st.columns([3, 2])
        with sec_col1:
            sec_rows = []
            for s, pct in sorted(sector_exposure.items(), key=lambda x: -x[1]):
                val = sum(h["value"] for h in holdings.values() if h["sector"] == s)
                count = sum(1 for h in holdings.values() if h["sector"] == s)
                sec_rows.append({
                    "Sector": SECTOR_LABELS.get(s, s),
                    "Holdings": count,
                    "Value": f"${val:,.2f}",
                    "Weight": f"{pct:.1f}%",
                })
            st.dataframe(pd.DataFrame(sec_rows), use_container_width=True, hide_index=True)

    # ==================================================================
    # SECTION 3: RECOMMENDED PORTFOLIO
    # ==================================================================
    st.markdown("<div class='section-header'>Recommended Portfolio</div>", unsafe_allow_html=True)
    st.markdown(f"**Advisor: {agent['name']}**")
    st.markdown(f"<div class='plan-step'>{agent['philosophy']}</div>", unsafe_allow_html=True)

    if profile_reasons:
        st.markdown("**Your allocation has been customized based on your profile:**")
        for r in profile_reasons:
            st.markdown(f"<div class='info-card'>{r}</div>", unsafe_allow_html=True)

    # Current vs Target bar chart
    bar_cats = [ASSET_CLASSES[c]["label"] for c in ASSET_CLASSES]
    bar_current = [current_pct_by_class.get(c, 0) for c in ASSET_CLASSES]
    bar_target = [target_pct_by_class.get(c, 0) for c in ASSET_CLASSES]
    fig_bar = dark_bar_chart(bar_cats, bar_current, bar_target, "Current vs. Target Allocation")
    st.pyplot(fig_bar)
    plt.close(fig_bar)

    # Target allocation table
    target_rows = []
    for cls in ASSET_CLASSES:
        t_pct = target_pct_by_class[cls]
        if t_pct <= 0:
            continue
        t_val = target_by_class[cls]
        c_val = current_by_class.get(cls, 0)
        c_pct = current_pct_by_class.get(cls, 0)
        diff = diff_by_class[cls]
        pick = agent["etf_picks"].get(cls, {})
        target_rows.append({
            "Asset Class": ASSET_CLASSES[cls]["label"],
            "Target %": f"{t_pct:.1f}%",
            "Target $": f"${t_val:,.2f}",
            "Current %": f"{c_pct:.1f}%",
            "Current $": f"${c_val:,.2f}",
            "Gap": f"${diff:+,.2f}",
            "Recommended ETF": f"{pick.get('ticker', '')} - {pick.get('name', '')}",
        })
    st.dataframe(pd.DataFrame(target_rows), use_container_width=True, hide_index=True)

    # ==================================================================
    # SECTION 4: YOUR REBALANCING PLAN
    # ==================================================================
    st.markdown("<div class='section-header'>Your Rebalancing Plan</div>", unsafe_allow_html=True)

    # Build trades
    trades = []

    # --- SELLS: reduce overweight asset classes ---
    for cls in ASSET_CLASSES:
        diff = diff_by_class[cls]
        if diff >= 0:
            continue
        sell_amount = abs(diff)
        # Find holdings in this class to sell, largest first
        class_holdings = sorted(
            [(t, h) for t, h in holdings.items() if h["asset_class"] == cls],
            key=lambda x: -x[1]["value"]
        )
        # Prefer selling non-recommended ETFs first
        rec_ticker = agent["etf_picks"].get(cls, {}).get("ticker", "")
        class_holdings.sort(key=lambda x: 0 if x[0] != rec_ticker else 1)

        remaining = sell_amount
        for t, h in class_holdings:
            if remaining <= 0:
                break
            sell_val = min(h["value"], remaining)
            sell_shares = round(sell_val / h["price"], 2) if h["price"] > 0 else 0
            reason = f"Reduce {ASSET_CLASSES[cls]['label']} allocation from {current_pct_by_class[cls]:.1f}% toward {target_pct_by_class[cls]:.1f}% target"
            if h["pct"] > 20:
                reason += f" (also reduces {t} concentration from {h['pct']:.1f}%)"
            trades.append({
                "action": "SELL", "ticker": t, "name": h["info"]["name"][:25],
                "shares": sell_shares, "amount": sell_val, "reason": reason,
                "asset_class": cls,
            })
            remaining -= sell_val

    # --- BUYS: fill underweight asset classes ---
    for cls in ASSET_CLASSES:
        diff = diff_by_class[cls]
        if diff <= 0:
            continue
        pick = agent["etf_picks"].get(cls)
        if not pick:
            continue
        buy_ticker = pick["ticker"]
        live_p = get_live_price(buy_ticker) or 0
        buy_shares = round(diff / live_p, 2) if live_p > 0 else 0
        trades.append({
            "action": "BUY", "ticker": buy_ticker, "name": pick["name"],
            "shares": buy_shares, "amount": diff, "reason": pick["why"],
            "asset_class": cls,
        })

    if not trades:
        st.markdown("<div class='hold-card'><strong>Your portfolio is well-aligned with your target strategy.</strong> No trades needed at this time.</div>", unsafe_allow_html=True)
    else:
        total_sells = sum(t["amount"] for t in trades if t["action"] == "SELL")
        total_buys = sum(t["amount"] for t in trades if t["action"] == "BUY")

        s1, s2 = st.columns(2)
        with s1:
            st.markdown(f"<div class='sell-card'><strong>Total to Sell:</strong> <span class='red'>${total_sells:,.2f}</span></div>", unsafe_allow_html=True)
        with s2:
            st.markdown(f"<div class='buy-card'><strong>Total to Buy:</strong> <span class='green'>${total_buys:,.2f}</span></div>", unsafe_allow_html=True)

        # Sells first
        sell_trades = [t for t in trades if t["action"] == "SELL"]
        buy_trades = [t for t in trades if t["action"] == "BUY"]

        if sell_trades:
            st.markdown("**Step 1: Sell overweight positions**")
            for t in sell_trades:
                shares_txt = f" (~{t['shares']} shares)" if t['shares'] > 0 else ""
                st.markdown(
                    f"<div class='sell-card'>"
                    f"<span class='red'><strong>SELL</strong></span> "
                    f"<strong>{t['ticker']}</strong> ({t['name']}) -- ${t['amount']:,.2f}{shares_txt}<br>"
                    f"<span class='muted'>{t['reason']}</span></div>",
                    unsafe_allow_html=True,
                )

        if buy_trades:
            st.markdown("**Step 2: Buy into target allocation**")
            for t in buy_trades:
                shares_txt = f" (~{t['shares']} shares)" if t['shares'] > 0 else ""
                st.markdown(
                    f"<div class='buy-card'>"
                    f"<span class='green'><strong>BUY</strong></span> "
                    f"<strong>{t['ticker']}</strong> ({t['name']}) -- ${t['amount']:,.2f}{shares_txt}<br>"
                    f"<span class='muted'>{t['reason']}</span></div>",
                    unsafe_allow_html=True,
                )

        # Plain English summary
        st.markdown("<div class='plan-step'>", unsafe_allow_html=True)
        st.markdown("**Summary**")

        equity_classes = ["us_large_growth", "us_broad", "us_dividend_value", "intl_developed", "intl_emerging"]
        bond_classes = ["us_core_bonds", "treasury_tips", "corporate_bonds"]
        target_eq = sum(target_pct_by_class.get(c, 0) for c in equity_classes)
        target_fi = sum(target_pct_by_class.get(c, 0) for c in bond_classes)

        summary_parts = []
        summary_parts.append(f"Your portfolio is currently **{total_stocks_pct:.0f}% equities / {total_bonds_pct:.0f}% fixed income**. "
                           f"Your personalized target is **{target_eq:.0f}% / {target_fi:.0f}%**.")

        # Key observations
        over_classes = [(c, current_pct_by_class[c] - target_pct_by_class[c]) for c in ASSET_CLASSES
                       if current_pct_by_class.get(c, 0) - target_pct_by_class.get(c, 0) > 3]
        under_classes = [(c, target_pct_by_class[c] - current_pct_by_class.get(c, 0)) for c in ASSET_CLASSES
                        if target_pct_by_class.get(c, 0) - current_pct_by_class.get(c, 0) > 3]

        if over_classes:
            over_str = ", ".join(f"{ASSET_CLASSES[c]['label']} (+{d:.0f}%)" for c, d in over_classes)
            summary_parts.append(f"You are **overweight** in: {over_str}.")
        if under_classes:
            under_str = ", ".join(f"{ASSET_CLASSES[c]['label']} (-{d:.0f}%)" for c, d in under_classes)
            summary_parts.append(f"You are **underweight** in: {under_str}.")

        summary_parts.append("Execute the sells first to free up capital, then place the buy orders. "
                           "This rebalances your portfolio toward your target allocation using diversified, low-cost ETFs.")

        for p in summary_parts:
            st.markdown(p)
        st.markdown("</div>", unsafe_allow_html=True)

    # ==================================================================
    # SECTION 5: EXCEL EXPORT
    # ==================================================================
    st.markdown("<div class='section-header'>Export Report</div>", unsafe_allow_html=True)

    if st.button("Generate Excel Report", use_container_width=True):
        wb = Workbook()
        wb.remove(wb.active)

        # Styles
        hdr_font = Font(bold=True, size=14, color="FFFFFF")
        hdr_fill = PatternFill(start_color="111111", end_color="111111", fill_type="solid")
        col_font = Font(bold=True, size=11, color="FFFFFF")
        col_fill = PatternFill(start_color="1e1e1e", end_color="1e1e1e", fill_type="solid")
        green_font = Font(color="00c805", bold=True)
        red_font = Font(color="ff5000", bold=True)
        muted_font = Font(color="8c8c8c", size=10)
        wrap_align = Alignment(wrap_text=True, vertical="top")
        money_fmt = '$#,##0.00'
        pct_fmt = '0.0"%"'

        def style_col_headers(ws, row, num_cols):
            for c in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=c)
                cell.font = col_font
                cell.fill = col_fill
                cell.alignment = Alignment(horizontal="center")

        def auto_width(ws, min_w=10, max_w=45):
            for col_cells in ws.columns:
                lengths = []
                for cell in col_cells:
                    if cell.value:
                        lengths.append(len(str(cell.value)))
                if lengths:
                    w = min(max(max(lengths) + 2, min_w), max_w)
                    ws.column_dimensions[get_column_letter(col_cells[0].column)].width = w

        # ---- TAB 1: CURRENT PORTFOLIO + ANALYSIS ----
        ws1 = wb.create_sheet("Current Portfolio")
        ws1.sheet_properties.tabColor = "3b82f6"
        ws1["A1"] = "CURRENT PORTFOLIO ANALYSIS"
        ws1["A1"].font = hdr_font
        ws1["A2"] = f"Generated {datetime.now().strftime('%B %d, %Y')}  |  Total Value: ${total_value:,.2f}  |  {len(holdings)} Holdings"
        ws1["A2"].font = muted_font

        h1 = ["Ticker", "Company", "Sector", "Asset Class", "Shares", "Price", "Value", "Weight %", "Dividend Yield", "P/E Ratio", "Alerts"]
        for c, h in enumerate(h1, 1):
            ws1.cell(row=4, column=c, value=h)
        style_col_headers(ws1, 4, len(h1))

        r = 5
        for t, h in sorted(holdings.items(), key=lambda x: -x[1]["value"]):
            ws1.cell(row=r, column=1, value=t)
            ws1.cell(row=r, column=2, value=h["info"]["name"][:40])
            ws1.cell(row=r, column=3, value=SECTOR_LABELS.get(h["sector"], h["sector"]))
            ws1.cell(row=r, column=4, value=ASSET_CLASSES.get(h["asset_class"], {}).get("label", ""))
            ws1.cell(row=r, column=5, value=round(h["shares"], 4))
            ws1.cell(row=r, column=6, value=h["price"]).number_format = money_fmt
            ws1.cell(row=r, column=7, value=h["value"]).number_format = money_fmt
            ws1.cell(row=r, column=8, value=round(h["pct"], 1))
            dy = h["info"].get("dividend_yield", 0)
            ws1.cell(row=r, column=9, value=f"{dy*100:.2f}%" if dy else "N/A")
            pe = h["info"].get("pe_ratio", 0)
            ws1.cell(row=r, column=10, value=f"{pe:.1f}" if pe else "N/A")
            alerts = []
            if h["pct"] > 20:
                alerts.append("HIGH CONCENTRATION")
            if h["sector"] not in ("etf", "unknown") and sector_exposure.get(h["sector"], 0) > 30:
                alerts.append("SECTOR OVERWEIGHT")
            alert_cell = ws1.cell(row=r, column=11, value=", ".join(alerts) if alerts else "")
            if alerts:
                alert_cell.font = red_font
            r += 1

        tr = r + 1
        ws1.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
        ws1.cell(row=tr, column=7, value=total_value).number_format = money_fmt
        ws1.cell(row=tr, column=7).font = Font(bold=True)

        # Allocation summary
        ar = tr + 2
        ws1.cell(row=ar, column=1, value="ALLOCATION BY ASSET CLASS").font = Font(bold=True, size=12)
        ws1.cell(row=ar+1, column=1, value="Asset Class")
        ws1.cell(row=ar+1, column=2, value="Current %")
        ws1.cell(row=ar+1, column=3, value="Current Value")
        style_col_headers(ws1, ar+1, 3)
        for i, cls in enumerate(ASSET_CLASSES):
            pct = current_pct_by_class.get(cls, 0)
            if pct > 0:
                ws1.cell(row=ar+2+i, column=1, value=ASSET_CLASSES[cls]["label"])
                ws1.cell(row=ar+2+i, column=2, value=round(pct, 1))
                ws1.cell(row=ar+2+i, column=3, value=current_by_class.get(cls, 0)).number_format = money_fmt

        auto_width(ws1)

        # ---- TAB 2: RECOMMENDED PORTFOLIO ----
        ws2 = wb.create_sheet("Recommended Portfolio")
        ws2.sheet_properties.tabColor = "00c805"
        ws2["A1"] = f"RECOMMENDED PORTFOLIO  —  {agent['name']}"
        ws2["A1"].font = hdr_font
        ws2["A2"] = agent["philosophy"]
        ws2["A2"].font = muted_font
        ws2["A2"].alignment = wrap_align

        h2 = ["Asset Class", "Target %", "Target Value", "Current %", "Current Value", "Gap ($)", "Recommended ETF", "ETF Name", "Rationale"]
        for c, h in enumerate(h2, 1):
            ws2.cell(row=4, column=c, value=h)
        style_col_headers(ws2, 4, len(h2))

        r = 5
        for cls in ASSET_CLASSES:
            tp = target_pct_by_class[cls]
            if tp <= 0:
                continue
            ws2.cell(row=r, column=1, value=ASSET_CLASSES[cls]["label"])
            ws2.cell(row=r, column=2, value=round(tp, 1))
            ws2.cell(row=r, column=3, value=target_by_class[cls]).number_format = money_fmt
            ws2.cell(row=r, column=4, value=round(current_pct_by_class.get(cls, 0), 1))
            ws2.cell(row=r, column=5, value=current_by_class.get(cls, 0)).number_format = money_fmt
            gap_cell = ws2.cell(row=r, column=6, value=diff_by_class[cls])
            gap_cell.number_format = money_fmt
            gap_cell.font = green_font if diff_by_class[cls] >= 0 else red_font
            pick = agent["etf_picks"].get(cls, {})
            ws2.cell(row=r, column=7, value=pick.get("ticker", ""))
            ws2.cell(row=r, column=8, value=pick.get("name", ""))
            ws2.cell(row=r, column=9, value=pick.get("why", ""))
            ws2.cell(row=r, column=9).alignment = wrap_align
            r += 1

        # Profile adjustments
        if profile_reasons:
            pr = r + 2
            ws2.cell(row=pr, column=1, value="PROFILE ADJUSTMENTS").font = Font(bold=True, size=12)
            for i, reason in enumerate(profile_reasons):
                ws2.cell(row=pr+1+i, column=1, value=reason)
                ws2.cell(row=pr+1+i, column=1).alignment = wrap_align

        auto_width(ws2)

        # ---- TAB 3: TRADING INSTRUCTIONS ----
        ws3 = wb.create_sheet("Trading Instructions")
        ws3.sheet_properties.tabColor = "f59e0b"
        ws3["A1"] = "STEP-BY-STEP TRADING INSTRUCTIONS"
        ws3["A1"].font = hdr_font
        ws3["A2"] = "Execute sells first to free up capital, then place buy orders."
        ws3["A2"].font = muted_font

        h3 = ["Step", "Action", "Ticker", "Name", "Asset Class", "Shares", "Amount ($)", "Reason"]
        for c, h in enumerate(h3, 1):
            ws3.cell(row=4, column=c, value=h)
        style_col_headers(ws3, 4, len(h3))

        for i, t in enumerate(trades, 5):
            ws3.cell(row=i, column=1, value=i - 4)
            a_cell = ws3.cell(row=i, column=2, value=t["action"])
            a_cell.font = red_font if t["action"] == "SELL" else green_font
            ws3.cell(row=i, column=3, value=t["ticker"])
            ws3.cell(row=i, column=4, value=t["name"])
            ws3.cell(row=i, column=5, value=ASSET_CLASSES.get(t["asset_class"], {}).get("label", ""))
            ws3.cell(row=i, column=6, value=t["shares"])
            ws3.cell(row=i, column=7, value=t["amount"]).number_format = money_fmt
            ws3.cell(row=i, column=8, value=t["reason"])
            ws3.cell(row=i, column=8).alignment = wrap_align

        sr = 5 + len(trades) + 2
        ws3.cell(row=sr, column=1, value="SUMMARY").font = Font(bold=True, size=12)
        ws3.cell(row=sr+1, column=1, value=f"Total Portfolio Value: ${total_value:,.2f}")
        ws3.cell(row=sr+2, column=1, value=f"Strategy: {agent['name']} ({risk_level})")
        ws3.cell(row=sr+3, column=1, value=f"Target: {sum(target_pct_by_class.get(c,0) for c in equity_classes):.0f}% Equities / {sum(target_pct_by_class.get(c,0) for c in bond_classes):.0f}% Fixed Income")
        ws3.cell(row=sr+4, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        auto_width(ws3)

        # ---- TAB 4: INVESTMENT PLAN ----
        ws4 = wb.create_sheet("Your Investment Plan")
        ws4.sheet_properties.tabColor = "8b5cf6"
        ws4["A1"] = "YOUR PERSONALIZED INVESTMENT PLAN"
        ws4["A1"].font = hdr_font

        lines = [
            "",
            "INVESTOR PROFILE",
            f"  Risk Tolerance:      {risk_level}",
            f"  Time Horizon:        {time_horizon}",
            f"  Primary Goal:        {goal}",
            f"  Age Range:           {age_range}",
            f"  International:       {intl_pref}",
            "",
            f"MATCHED STRATEGY: {agent['name'].upper()}",
            agent["philosophy"],
            "",
            "CURRENT STATE",
            f"  Total portfolio value: ${total_value:,.2f}",
            f"  Current allocation: {total_stocks_pct:.0f}% equities, {total_bonds_pct:.0f}% fixed income",
            f"  Target allocation:  {target_eq:.0f}% equities, {target_fi:.0f}% fixed income",
            f"  Number of holdings: {len(holdings)}",
            "",
        ]

        if profile_reasons:
            lines.append("PROFILE-BASED ADJUSTMENTS")
            for r in profile_reasons:
                lines.append(f"  - {r}")
            lines.append("")

        if warnings:
            lines.append("KEY RISKS IDENTIFIED")
            for w in warnings:
                lines.append(f"  - {w.replace('**', '')}")
            lines.append("")

        lines.append("RECOMMENDED ACTIONS")
        for i, t in enumerate(trades, 1):
            action = "Sell" if t["action"] == "SELL" else "Buy"
            lines.append(f"  {i}. {action} {t['ticker']} ({t['name']}) — ${t['amount']:,.2f} (~{t['shares']} shares)")
            lines.append(f"     Reason: {t['reason']}")
        lines.append("")

        lines += [
            "KEY PRINCIPLES",
            "  1. Diversification: Spread risk across asset classes, sectors, and geographies.",
            "  2. Low costs: Recommended ETFs have expense ratios under 0.10% in most cases.",
            "  3. Rebalance periodically: Review your allocation every 6-12 months.",
            "  4. Stay disciplined: Avoid emotional reactions to short-term market movements.",
            "  5. Tax awareness: Consider tax-loss harvesting when selling at a loss.",
            "",
            "DISCLAIMER",
            "  This report is for educational and informational purposes only.",
            "  It does not constitute personalized financial advice. Past performance",
            "  does not guarantee future results. Consult a licensed financial advisor",
            "  before making investment decisions.",
        ]

        for i, line in enumerate(lines, 3):
            cell = ws4.cell(row=i, column=1, value=line)
            if line.isupper() and line.strip():
                cell.font = Font(bold=True, size=12)
        ws4.column_dimensions["A"].width = 95

        # Save
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        st.download_button(
            label="Download Excel Report",
            data=buf.getvalue(),
            file_name=f"Portfolio_Rebalance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.success("Report generated successfully.")

else:
    st.markdown(
        "<div style='text-align:center; padding: 100px 20px;'>"
        "<h2 style='color:#ffffff; font-weight:700;'>Get Started</h2>"
        "<p style='color:#6e6e6e; font-size:1.05rem; max-width:460px; margin:0 auto;'>"
        "Answer a few questions about your investment preferences, "
        "then enter your portfolio holdings in the sidebar.</p>"
        "</div>",
        unsafe_allow_html=True,
    )
